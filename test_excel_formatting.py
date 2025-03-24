import unittest
import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def ensure_argb_format(color_str):
    """Ensure color is in aRGB format (AARRGGBB)"""
    if not color_str:
        return None
        
    # Strip any non-hex characters
    color_str = ''.join(c for c in color_str if c in '0123456789ABCDEFabcdef')
    
    # Handle different formats
    if len(color_str) == 6:  # RRGGBB -> FFRRGGBB
        return 'FF' + color_str
    elif len(color_str) == 8:  # Already AARRGGBB
        return color_str
    else:
        print(f"Warning: Unexpected color format: {color_str}")
        return 'FF000000'  # Default to black

def extract_cell_formatting(cell):
    """Extract formatting details from a cell for storage."""
    # Handle font properties
    font_props = {
        "name": cell.font.name,
        "size": cell.font.size,
        "bold": cell.font.bold,
        "italic": cell.font.italic,
        "underline": cell.font.underline,
    }
    
    # Extract font color
    if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
        font_props["color"] = ensure_argb_format(str(cell.font.color.rgb))
    else:
        font_props["color"] = None
    
    # Handle fill properties
    fill_props = {
        "fill_type": cell.fill.fill_type
    }
    
    # Extract fill color
    if cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
        fill_props["start_color"] = ensure_argb_format(str(cell.fill.start_color.rgb))
    else:
        fill_props["start_color"] = None
    
    # Handle border properties
    border_props = {}
    for side in ['top', 'bottom', 'left', 'right']:
        side_obj = getattr(cell.border, side)
        if not side_obj:
            border_props[side] = {"style": None, "color": None}
            continue
        
        # Extract border color
        if side_obj.color and hasattr(side_obj.color, 'rgb') and side_obj.color.rgb:
            side_color = ensure_argb_format(str(side_obj.color.rgb))
        else:
            side_color = None
        
        border_props[side] = {
            "style": side_obj.style,
            "color": side_color
        }
    
    return {
        "number_format": cell.number_format,
        "font": font_props,
        "fill": fill_props,
        "border": border_props,
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
        }
    }

def apply_formatting(cell, fmt_dict):
    """Apply formatting details to the cell."""
    try:
        # Apply number format
        cell.number_format = fmt_dict["number_format"]
        
        # Apply font formatting
        font_data = fmt_dict["font"]
        font_color = font_data.get("color")
        
        cell.font = Font(
            name=font_data.get("name"),
            size=font_data.get("size"),
            bold=font_data.get("bold"),
            italic=font_data.get("italic"),
            underline=font_data.get("underline"),
            color=font_color
        )
        
        # Apply fill formatting
        fill_data = fmt_dict["fill"]
        fill_type = fill_data.get("fill_type")
        
        if fill_type and fill_type not in ['none', None]:
            start_color = fill_data.get("start_color")
            if start_color:
                cell.fill = PatternFill(
                    fill_type=fill_type,
                    start_color=start_color,
                    end_color=start_color
                )
        
        # Apply alignment
        align_data = fmt_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=align_data.get("horizontal"),
            vertical=align_data.get("vertical"),
            wrap_text=align_data.get("wrap_text", False)
        )
        
        # Apply borders
        border_data = fmt_dict["border"]
        sides = {}
        
        for side_name in ['top', 'bottom', 'left', 'right']:
            side_data = border_data.get(side_name, {})
            style = side_data.get("style")
            color = side_data.get("color")
            
            if style:
                sides[side_name] = Side(style=style, color=color)
            else:
                sides[side_name] = Side()
        
        cell.border = Border(**sides)
            
    except Exception as e:
        print(f"Error applying formatting to cell {cell.coordinate}: {e}")

def fix_font_colors_only(file, output_file=None):
    """Fix ONLY font colors in workbook while preserving all other formatting."""
    if not os.path.exists(file):
        print(f"File not found: {file}")
        return None
    
    print(f"Processing file: {file}")
    
    # Create output filename if not provided
    if not output_file:
        base_name, ext = os.path.splitext(file)
        output_file = f"{base_name}_fixed{ext}"
    
    # Load the workbook
    wb = load_workbook(file)
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Process all cells
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                # Skip empty cells
                if cell.value is None:
                    continue
                
                # Extract all formatting
                fmt = extract_cell_formatting(cell)
                
                # Only modify the font color, keeping everything else the same
                if fmt["font"]["color"]:
                    fmt["font"]["color"] = "FF000000"  # Solid black
                
                # Apply the modified formatting back
                apply_formatting(cell, fmt)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Saved fixed file: {output_file}")
    return output_file

class TestExcelFormatting(unittest.TestCase):
    """Test if cell formatting is preserved correctly in Excel files."""
    
    @classmethod
    def setUpClass(cls):
        """Create a test Excel file with various formatting."""
        cls.test_dir = "format_test"
        os.makedirs(cls.test_dir, exist_ok=True)
        cls.source_file = os.path.join(cls.test_dir, "source_formatting.xlsx")
        cls.recreated_file = os.path.join(cls.test_dir, "recreated_formatting.xlsx")
        
        # Create source workbook with formatted cells
        wb = Workbook()
        ws = wb.active
        ws.title = "FormattingTest"
        
        # Cell with font formatting - RED
        ws['A1'] = "Red Font"
        ws['A1'].font = Font(color="FF0000", bold=True, size=14)  # Full red
        
        # Cell with fill formatting - GREEN
        ws['A2'] = "Green Fill"
        ws['A2'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Full green
        
        # Cell with border formatting - BLUE
        ws['A3'] = "Blue Border"
        ws['A3'].border = Border(
            left=Side(style='thin', color="0000FF"),  # Full blue
            right=Side(style='thin', color="0000FF"),
            top=Side(style='thin', color="0000FF"),
            bottom=Side(style='thin', color="0000FF")
        )
        
        # Cell with multiple formats
        ws['A4'] = "Combined Formatting"
        ws['A4'].font = Font(color="FF00FF", italic=True)  # Magenta
        ws['A4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        ws['A4'].border = Border(
            left=Side(style='medium', color="00FFFF"),  # Cyan
            right=Side(style='medium', color="00FFFF"),
            top=Side(style='medium', color="00FFFF"),
            bottom=Side(style='medium', color="00FFFF")
        )
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Save the source workbook
        wb.save(cls.source_file)
        print(f"Created test source file: {cls.source_file}")
    
    def test_preservation_of_formatting(self):
        """Test if formatting is preserved when extracted and reapplied."""
        # Load the source workbook
        source_wb = load_workbook(self.source_file)
        source_ws = source_wb.active
        
        # Create a new workbook
        new_wb = Workbook()
        new_ws = new_wb.active
        
        # Copy values and formatting for each cell
        for coord in ['A1', 'A2', 'A3', 'A4']:
            # Copy the value
            new_ws[coord] = source_ws[coord].value
            
            # Extract formatting
            fmt_dict = extract_cell_formatting(source_ws[coord])
            
            # Print for debug
            print(f"\nCell {coord} formatting:")
            for key, value in fmt_dict.items():
                if key == 'font' and 'color' in value:
                    print(f"  font color: {value['color']}")
                if key == 'fill' and 'start_color' in value:
                    print(f"  fill color: {value['start_color']}")
                if key == 'border':
                    for side, side_data in value.items():
                        if 'color' in side_data and side_data['color']:
                            print(f"  border {side} color: {side_data['color']}")
            
            # Apply formatting
            apply_formatting(new_ws[coord], fmt_dict)
        
        # Save the recreated workbook
        new_wb.save(self.recreated_file)
        print(f"Created recreated file: {self.recreated_file}")
        
        # Reload both workbooks for comparison
        source_wb = load_workbook(self.source_file)
        source_ws = source_wb.active
        
        recreated_wb = load_workbook(self.recreated_file)
        recreated_ws = recreated_wb.active
        
        # Compare formatting with specific checks for each cell
        
        # A1: Red Font - check font color
        source_color = str(source_ws['A1'].font.color.rgb)
        recreated_color = str(recreated_ws['A1'].font.color.rgb)
        self.assertEqual(source_color, recreated_color, 
                         f"Font color in A1 not preserved: {source_color} vs {recreated_color}")
        
        # A2: Green Fill - check fill color
        source_fill = str(source_ws['A2'].fill.start_color.rgb)
        recreated_fill = str(recreated_ws['A2'].fill.start_color.rgb)
        self.assertEqual(source_fill, recreated_fill, 
                         f"Fill color in A2 not preserved: {source_fill} vs {recreated_fill}")
        
        # A3: Blue Border - check border color
        source_border = str(source_ws['A3'].border.top.color.rgb)
        recreated_border = str(recreated_ws['A3'].border.top.color.rgb)
        self.assertEqual(source_border, recreated_border, 
                         f"Border color in A3 not preserved: {source_border} vs {recreated_border}")
        
        # A4: Combined formatting - check multiple aspects
        self.assertEqual(str(source_ws['A4'].font.color.rgb), str(recreated_ws['A4'].font.color.rgb),
                         "Font color in A4 not preserved")
        self.assertEqual(str(source_ws['A4'].fill.start_color.rgb), str(recreated_ws['A4'].fill.start_color.rgb),
                         "Fill color in A4 not preserved")
        self.assertEqual(str(source_ws['A4'].border.top.color.rgb), str(recreated_ws['A4'].border.top.color.rgb),
                         "Border color in A4 not preserved")
        self.assertEqual(source_ws['A4'].alignment.horizontal, recreated_ws['A4'].alignment.horizontal,
                         "Horizontal alignment in A4 not preserved")
    
    def test_font_color_fixing(self):
        """Test that ONLY font colors are changed when fixing."""
        # Use the recreated file from the previous test
        if not os.path.exists(self.recreated_file):
            self.test_preservation_of_formatting()
        
        # Apply font fixing
        fixed_file = fix_font_colors_only(self.recreated_file)
        
        # The fixed file should exist
        self.assertTrue(os.path.exists(fixed_file), "Fixed file was not created")
        
        # Load source and fixed files
        source_wb = load_workbook(self.source_file)
        source_ws = source_wb.active
        
        fixed_wb = load_workbook(fixed_file)
        fixed_ws = fixed_wb.active
        
        # All font colors should be black
        for coord in ['A1', 'A2', 'A3', 'A4']:
            if fixed_ws[coord].font.color:
                fixed_color = str(fixed_ws[coord].font.color.rgb)
                self.assertEqual(fixed_color, "FF000000", 
                                f"Font in {coord} is not black: {fixed_color}")
        
        # Other formatting should be preserved
        
        # A2: Green Fill - check fill color is preserved
        source_fill = str(source_ws['A2'].fill.start_color.rgb)
        fixed_fill = str(fixed_ws['A2'].fill.start_color.rgb)
        self.assertEqual(source_fill, fixed_fill, 
                        f"Fill color in A2 not preserved: {source_fill} vs {fixed_fill}")
        
        # A3: Blue Border - check border color is preserved
        source_border = str(source_ws['A3'].border.top.color.rgb)
        fixed_border = str(fixed_ws['A3'].border.top.color.rgb)
        self.assertEqual(source_border, fixed_border, 
                        f"Border color in A3 not preserved: {source_border} vs {fixed_border}")
        
        # A4: Combined formatting - check non-font formatting is preserved
        source_fill = str(source_ws['A4'].fill.start_color.rgb)
        fixed_fill = str(fixed_ws['A4'].fill.start_color.rgb)
        self.assertEqual(source_fill, fixed_fill, 
                        f"Fill color in A4 not preserved: {source_fill} vs {fixed_fill}")
        
        source_border = str(source_ws['A4'].border.top.color.rgb)
        fixed_border = str(fixed_ws['A4'].border.top.color.rgb)
        self.assertEqual(source_border, fixed_border, 
                        f"Border color in A4 not preserved: {source_border} vs {fixed_border}")
        
        self.assertEqual(source_ws['A4'].alignment.horizontal, fixed_ws['A4'].alignment.horizontal,
                        "Alignment in A4 not preserved")

if __name__ == "__main__":
    unittest.main(verbosity=2)