import pandas as pd
import sqlite3
import json
import os
import re
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from copy import copy

# ============================================================================
# Configuration
# ============================================================================
excel_files = ["Deposits Data Lite.xlsx", "Form X Report  Main Lite.xlsx", "Loans Data Lite.xlsx"]
report_sheets = {"Part I", "Part II", "Part III", "MIS-Report"}
exclude_sheets = {"Pivot-Borrowings"}
db_filename = "excel_data.db"
output_dir = "output"
new_base_path = ""  # Will be set at runtime

# ============================================================================
# Utility Functions
# ============================================================================
class DateTimeEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle datetime objects."""
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super(DateTimeEncoder, self).default(obj)

def create_excel_file_map(excel_files):
    """Create a mapping dictionary for Excel filenames with variations."""
    excel_file_map = {}
    for file in excel_files:
        base_name = os.path.basename(file)
        name_without_ext = os.path.splitext(base_name)[0]
        clean_name = base_name.replace(" ", "")
        clean_name_without_ext = name_without_ext.replace(" ", "")
        
        # Add all variations to the map
        for key in [base_name, name_without_ext, clean_name, clean_name_without_ext]:
            excel_file_map[key] = base_name
    return excel_file_map

def copy_cell_formatting(source_cell, target_cell):
    """Copy all formatting properties from source cell to target cell."""
    try:
        # Copy font properties
        target_cell.font = copy(source_cell.font)
        
        # Copy fill properties
        target_cell.fill = copy(source_cell.fill)
        
        # Copy border properties
        target_cell.border = copy(source_cell.border)
        
        # Copy number format
        target_cell.number_format = source_cell.number_format
        
        # Copy protection
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        
        # Copy alignment
        target_cell.alignment = copy(source_cell.alignment)
        
    except Exception as e:
        print(f"Error copying formatting from {source_cell.coordinate}: {e}")

# ============================================================================
# Formula Path Handling
# ============================================================================
def fix_external_references(formula, excel_file_map):
    """
    Fix external references in Excel formulas.
    """
    if not formula or not isinstance(formula, str):
        return formula
    
    # Log the formula for debugging
    if "xlsx" in formula or "[" in formula and "]" in formula:
        print(f"Analyzing formula with potential external reference: {formula}")
    
    # Pattern to match standard external references like: 'Path\[Filename.xlsx]SheetName'!Range
    standard_pattern = r"'?([^']*\[([^]]+)\]([^!']*))'?!([A-Z0-9:$]+)"
    
    # Pattern to match references with numeric workbook indices like: [1]Deposits!L:L
    indexed_pattern = r"\[(\d+)\]([^!]+)!([A-Z0-9:$]+)"
    
    # Pattern to match non-conventional references that might include sheet names or file paths
    sheet_reference_pattern = r"([\w\s-]+\.xlsx)([\w\s-]+)"
    
    def replace_standard_match(match):
        full_path = match.group(1)  # The whole path with filename and sheet
        filename = match.group(2)   # Just the filename
        sheet = match.group(3)      # Just the sheet name
        cell_ref = match.group(4)   # The cell reference
        
        print(f"Found standard external reference - File: {filename}, Sheet: {sheet}, Cell: {cell_ref}")
        
        # Check if this filename matches any of our known files (possibly with slight differences)
        target_filename = None
        for file_key, file_value in excel_file_map.items():
            # Case-insensitive comparison, ignore spaces and parentheses
            clean_filename = filename.lower().replace(" ", "").replace("(", "").replace(")", "")
            clean_key = file_key.lower().replace(" ", "").replace("(", "").replace(")", "")
            if clean_filename in clean_key or clean_key in clean_filename:
                target_filename = file_value
                break
        
        if target_filename:
            # Construct the new reference with the new path and matched filename
            new_ref = f"'{new_base_path}[{target_filename}]{sheet}'!{cell_ref}"
            print(f"Updated to: {new_ref}")
            return new_ref
        else:
            # If no match found, keep the original reference
            return match.group(0)
    
    def replace_indexed_match(match):
        """Handle references with numeric workbook indices like [1]Deposits!L:L"""
        workbook_index = match.group(1)  # The numeric index (e.g., '1')
        sheet_name = match.group(2)      # The sheet name (e.g., 'Deposits')
        cell_ref = match.group(3)        # The range reference (e.g., 'L:L')
        
        print(f"Found indexed external reference - Index: [{workbook_index}], Sheet: {sheet_name}, Cell: {cell_ref}")
        
        # Map workbook indices to file names - customize based on your specific workbooks
        index_to_file = {
            '1': 'Deposits Data Lite.xlsx',
            '2': 'Loans Data Lite.xlsx',
            '3': 'Form X Report  Main Lite.xlsx'
        }
        
        if workbook_index in index_to_file and new_base_path:
            # For new base path, create a full path reference
            filename = index_to_file[workbook_index]
            new_ref = f"'{new_base_path}[{filename}]{sheet_name}'!{cell_ref}"
            print(f"Mapped indexed reference to: {new_ref}")
            return new_ref
        else:
            # Keep the original indexed reference if no base path change or unknown index
            print(f"Preserving original indexed reference: [{workbook_index}]{sheet_name}!{cell_ref}")
            return match.group(0)
    
    def replace_sheet_reference(match):
        file = match.group(1)
        sheet = match.group(2)
        print(f"Found non-standard reference - File: {file}, Content: {sheet}")
        
        # Try to match with our known files
        target_filename = None
        for file_key, file_value in excel_file_map.items():
            clean_file = file.lower().replace(" ", "").replace("(", "").replace(")", "")
            clean_key = file_key.lower().replace(" ", "").replace("(", "").replace(")", "")
            if clean_file in clean_key or clean_key in clean_file:
                target_filename = file_value
                break
        
        if target_filename and new_base_path:
            return f"{new_base_path}{target_filename}{sheet}"
        return match.group(0)
    
    # Apply different patterns in sequence
    updated_formula = re.sub(standard_pattern, replace_standard_match, formula)
    updated_formula = re.sub(indexed_pattern, replace_indexed_match, updated_formula)
    updated_formula = re.sub(sheet_reference_pattern, replace_sheet_reference, updated_formula)
    
    # Debug output when we change formulas
    if updated_formula != formula:
        print(f"Formula updated:\nFrom: {formula}\nTo:   {updated_formula}")
        
    return updated_formula

# ============================================================================
# Database Helper Functions
# ============================================================================
def setup_database(db_path):
    """Create database and necessary tables."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    print(f"Creating database schema in: {db_path}")
    
    # Create workbooks table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS workbooks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT UNIQUE,
        properties TEXT
    )
    """)
    
    # Create sheets table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS sheets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        workbook_id INTEGER,
        sheet_name TEXT,
        sheet_type TEXT,
        max_row INTEGER,
        max_column INTEGER,
        merged_cells TEXT,
        column_dimensions TEXT,
        row_dimensions TEXT,
        FOREIGN KEY (workbook_id) REFERENCES workbooks (id),
        UNIQUE (workbook_id, sheet_name)
    )
    """)
    
    # Create cells table - simplified to only store value and formula status
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS cells (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sheet_id INTEGER,
        coordinate TEXT,
        value TEXT,
        is_formula BOOLEAN,
        FOREIGN KEY (sheet_id) REFERENCES sheets (id),
        UNIQUE (sheet_id, coordinate)
    )
    """)
    
    # Create tabular_data table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tabular_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        workbook TEXT,
        sheet TEXT,
        table_name TEXT UNIQUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)
    
    conn.commit()
    return conn

def insert_workbook(cursor, filename, properties):
    """Insert workbook data into the database and return its ID."""
    cursor.execute(
        "INSERT OR REPLACE INTO workbooks (filename, properties) VALUES (?, ?)",
        (filename, json.dumps(properties))
    )
    cursor.execute("SELECT id FROM workbooks WHERE filename = ?", (filename,))
    return cursor.fetchone()[0]

def insert_sheet(cursor, workbook_id, sheet_name, sheet_type, max_row, max_column, merged_cells, column_dimensions, row_dimensions):
    """Insert sheet data into the database and return its ID."""
    cursor.execute(
        """INSERT OR REPLACE INTO sheets 
           (workbook_id, sheet_name, sheet_type, max_row, max_column, merged_cells, column_dimensions, row_dimensions) 
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (workbook_id, sheet_name, sheet_type, max_row, max_column, 
         json.dumps(merged_cells),
         json.dumps(column_dimensions),
         json.dumps(row_dimensions))
    )
    cursor.execute("SELECT id FROM sheets WHERE workbook_id = ? AND sheet_name = ?", 
                   (workbook_id, sheet_name))
    return cursor.fetchone()[0]

def insert_cell(cursor, sheet_id, coordinate, value, is_formula):
    """Insert cell data into the database (without formatting info)."""
    cursor.execute(
        "INSERT OR REPLACE INTO cells (sheet_id, coordinate, value, is_formula) VALUES (?, ?, ?, ?)",
        (sheet_id, coordinate, value, is_formula)
    )

# ============================================================================
# Phase 1: Data Identification
# ============================================================================
def identify_data():
    """Extract and identify data from Excel files."""
    print("\n" + "="*70)
    print("PHASE 1: DATA IDENTIFICATION")
    print("="*70)
    
    # Access global configuration variables
    global excel_files, report_sheets, exclude_sheets
    
    # Dictionary to store all workbook data and metadata
    workbook_data = {}
    
    # Create a list to store potential references for later analysis
    potential_references = []
    
    # Process each Excel file
    for file in excel_files:
        print(f"Processing file: {file}")
        workbook_data[file] = {"sheets": {}}
        
        # Load the workbook with openpyxl (data_only=False to capture formulas)
        wb = load_workbook(file, data_only=False)
        
        # Get workbook-level properties
        workbook_data[file]["properties"] = {
            "title": wb.properties.title,
            "creator": wb.properties.creator,
            "created": str(wb.properties.created) if wb.properties.created else None,
            "sheet_names": wb.sheetnames
        }
        
        # Process each sheet in the workbook
        for sheet_name in wb.sheetnames:
            # Skip excluded sheets
            if sheet_name in exclude_sheets:
                print(f"Skipping excluded sheet: {sheet_name}")
                continue
            
            # Get the worksheet
            ws = wb[sheet_name]
            
            # Initialize sheet data
            sheet_data = {
                "type": "report" if sheet_name in report_sheets else "non_report",
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_cells": [str(merged_range) for merged_range in ws.merged_cells.ranges],
                "column_dimensions": {col: {"width": ws.column_dimensions[col].width} 
                                     for col in ws.column_dimensions},
                "row_dimensions": {row: {"height": ws.row_dimensions[row].height} 
                                  for row in ws.row_dimensions},
                "cells": {}
            }
            
            # Process all cells in the sheet
            print(f"  Processing cells in sheet: {sheet_name}")
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Skip empty cells to save space
                    if cell.value is None:
                        continue
                    
                    # Enhanced detection for references and formulas
                    cell_value = cell.value
                    is_formula = False
                    
                    if isinstance(cell_value, str):
                        # Check for standard formula
                        if cell_value.startswith('='):
                            is_formula = True
                        
                        # Check for potential external references
                        if '.xlsx' in cell_value or '.xls' in cell_value or ('[' in cell_value and ']' in cell_value):
                            potential_references.append({
                                'file': file,
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'value': cell_value
                            })
                    
                    # Handle datetime objects for JSON serialization
                    if isinstance(cell_value, (datetime, date)):
                        cell_value = cell_value.isoformat()
                    
                    # Store cell data - only value and formula status
                    sheet_data["cells"][cell.coordinate] = {
                        "value": cell_value,
                        "is_formula": is_formula
                    }
            
            # Store sheet data in the workbook dictionary
            workbook_data[file]["sheets"][sheet_name] = sheet_data
        
        print(f"Completed processing file: {file}")
    
    # Output summary
    print("\nData Identification Summary:")
    for file, data in workbook_data.items():
        print(f"\nFile: {file}")
        print(f"  Total sheets: {len(data['sheets'])}")
        for sheet_name, sheet_data in data['sheets'].items():
            cell_count = len(sheet_data['cells'])
            sheet_type = sheet_data['type']
            print(f"  Sheet: {sheet_name} ({sheet_type}) - {cell_count} non-empty cells")
    
    # Save the identification results to a JSON file for reference
    with open('workbook_identification.json', 'w') as f:
        json.dump(workbook_data, f, indent=2, cls=DateTimeEncoder)
    
    return workbook_data

# ============================================================================
# Phase 2: Data Storage
# ============================================================================
def store_data(workbook_data=None):
    """Store identified data in SQLite database."""
    print("\n" + "="*70)
    print("PHASE 2: DATA STORAGE")
    print("="*70)
    
    # Access global configuration variables
    global excel_files, report_sheets, exclude_sheets, db_filename, new_base_path
    
    # Create mapping dictionary for Excel filenames
    excel_file_map = create_excel_file_map(excel_files)
    
    # Set up database - ensure we start with a clean database
    if os.path.exists(db_filename):
        try:
            os.remove(db_filename)
            print(f"Removed existing database: {db_filename}")
        except PermissionError:
            print(f"Could not remove existing database. Make sure it's not in use by another program.")
            raise
    
    conn = setup_database(db_filename)
    cursor = conn.cursor()
    
    # Process each Excel file
    for file in excel_files:
        print(f"Processing file: {file}")
        
        # Insert workbook metadata
        wb = load_workbook(file, data_only=False)
        properties = {
            "title": wb.properties.title,
            "creator": wb.properties.creator,
            "created": str(wb.properties.created) if wb.properties.created else None,
            "sheet_names": wb.sheetnames
        }
        
        workbook_id = insert_workbook(cursor, file, properties)
        conn.commit()
        
        # Process each sheet in the workbook
        for sheet_name in wb.sheetnames:
            if sheet_name in exclude_sheets:
                print(f"  Skipping excluded sheet: {sheet_name}")
                continue
            
            ws = wb[sheet_name]
            sheet_type = "report" if sheet_name in report_sheets else "non_report"
            
            # Store sheet metadata
            merged_cells = [str(merged_range) for merged_range in ws.merged_cells.ranges]
            column_dimensions = {col: {"width": ws.column_dimensions[col].width} 
                                for col in ws.column_dimensions}
            row_dimensions = {row: {"height": ws.row_dimensions[row].height}
                              for row in ws.row_dimensions}
            
            sheet_id = insert_sheet(cursor, workbook_id, sheet_name, sheet_type, ws.max_row, ws.max_column, 
                                    merged_cells, column_dimensions, row_dimensions)
            conn.commit()
            
            # Store cell data only (value and formula status)
            print(f"  Processing cells in sheet: {sheet_name}")
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Skip empty cells to save space
                    if cell.value is None:
                        continue
                    
                    # Determine if the cell contains a formula
                    is_formula = isinstance(cell.value, str) and cell.value.startswith('=')
                    
                    # Convert cell value to string for storage
                    if isinstance(cell.value, (int, float, bool)):
                        cell_value = str(cell.value)
                    elif cell.value is None:
                        cell_value = ""
                    else:
                        cell_value = str(cell.value)
                    
                    # Fix external references in formulas if needed
                    if is_formula and new_base_path:
                        cell_value = fix_external_references(cell_value, excel_file_map)
                    
                    # Store cell data (without formatting)
                    insert_cell(cursor, sheet_id, cell.coordinate, cell_value, is_formula)
            
            # For non-report sheets, also store as tabular data
            if sheet_type == "non_report":
                try:
                    base_name = os.path.splitext(os.path.basename(file))[0]
                    table_name = f"{base_name}_{sheet_name}".replace(" ", "_").replace("-", "_")
                    
                    # Read the sheet as a DataFrame
                    df = pd.read_excel(file, sheet_name=sheet_name)
                    
                    # Store the DataFrame in the database
                    df.to_sql(table_name, conn, if_exists='replace', index=False)
                    
                    # Record this table in the tabular_data table
                    cursor.execute(
                        "INSERT OR REPLACE INTO tabular_data (workbook, sheet, table_name) VALUES (?, ?, ?)",
                        (file, sheet_name, table_name)
                    )
                    
                    print(f"  Stored tabular data for sheet '{sheet_name}' in table '{table_name}'")
                except Exception as e:
                    print(f"  Error storing tabular data for sheet '{sheet_name}': {e}")
            
            conn.commit()
        
        print(f"Completed processing file: {file}")
    
    # Close the database connection
    conn.commit()
    conn.close()
    print("\nData storage complete.")

# ============================================================================
# Phase 3: Data Recreation
# ============================================================================
def recreate_workbooks():
    """Recreate workbooks from stored data with direct formatting copy."""
    print("\n" + "="*70)
    print("PHASE 3: DATA RECREATION WITH DIRECT FORMATTING COPY")
    print("="*70)
    
    # Access global configuration variables
    global excel_files, exclude_sheets, db_filename, output_dir, new_base_path
    
    # Create mapping dictionary for Excel filenames
    excel_file_map = create_excel_file_map(excel_files)
    
    # Create a mapping for numeric index references
    index_to_filename = {
        '1': 'Deposits Data Lite.xlsx',
        '2': 'Loans Data Lite.xlsx',
        '3': 'Form X Report  Main Lite.xlsx'
    }
    
    # Load all source workbooks into memory for direct formatting access
    source_workbooks = {}
    for file in excel_files:
        print(f"Loading source workbook: {file}")
        source_workbooks[file] = load_workbook(file, data_only=False)
    
    # Open database connection
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    
    # Get recreated workbook paths for later use in font fixing
    recreated_files = []
    
    # Process each workbook for recreation
    for file in excel_files:
        print(f"Recreating workbook: {file}")
        base_name = os.path.splitext(os.path.basename(file))[0]
        
        # Get workbook ID
        cursor.execute("SELECT id FROM workbooks WHERE filename = ?", (file,))
        workbook_id_result = cursor.fetchone()
        
        if not workbook_id_result:
            print(f"Workbook {file} not found in database. Skipping.")
            continue
        
        workbook_id = workbook_id_result[0]
        
        # Get sheet information for this workbook
        cursor.execute("""
            SELECT id, sheet_name, sheet_type, max_row, max_column, merged_cells, column_dimensions, row_dimensions
            FROM sheets
            WHERE workbook_id = ?
        """, (workbook_id,))
        sheets_info = cursor.fetchall()
        
        # Create a new workbook
        new_wb = Workbook()
        # Remove the default sheet
        if len(sheets_info) > 0:
            default_sheet = new_wb.active
            new_wb.remove(default_sheet)
        
        # Process each sheet
        for sheet_info in sheets_info:
            sheet_id, sheet_name, sheet_type, max_row, max_column, merged_cells, column_dimensions, row_dimensions = sheet_info
            
            print(f"  Recreating sheet: {sheet_name} (type: {sheet_type})")
            
            # Create the sheet
            new_ws = new_wb.create_sheet(title=sheet_name)
            
            # Get the source worksheet for direct formatting
            source_ws = source_workbooks[file][sheet_name]
            
            # Apply merged cells
            if merged_cells:
                for merged_range in json.loads(merged_cells):
                    new_ws.merge_cells(merged_range)
            
            # Apply dimensions
            for col_key, properties in json.loads(column_dimensions).items():
                if col_key in new_ws.column_dimensions and properties.get("width"):
                    new_ws.column_dimensions[col_key].width = properties["width"]
            
            for row_key, properties in json.loads(row_dimensions).items():
                row = int(row_key)
                if row in new_ws.row_dimensions and properties.get("height"):
                    new_ws.row_dimensions[row].height = properties["height"]
            
            # Populate sheet data
            cursor.execute("""
                SELECT coordinate, value, is_formula
                FROM cells
                WHERE sheet_id = ?
            """, (sheet_id,))
            cells_data = cursor.fetchall()
            
            for cell_data in cells_data:
                coordinate, value, is_formula = cell_data
                
                # Set the cell value
                if is_formula:
                    # Handle formula cells
                    if new_base_path:
                        formula_value = fix_external_references(value, excel_file_map)
                    else:
                        formula_value = value
                    
                    # Special handling for indexed references in specific sheets
                    is_special_sheet = sheet_name in ["MIS-Report", "Part I", "Part II", "Part III"]
                    has_indexed_ref = ('[1]' in value) or ('[2]' in value) or ('[3]' in value)
                    
                    try:
                        if is_special_sheet and has_indexed_ref:
                            # For indexed references, set value directly
                            new_ws[coordinate].value = formula_value
                        elif formula_value.startswith('='):
                            # Standard formula
                            new_ws[coordinate].value = None
                            new_ws[coordinate].formula = formula_value[1:]
                        else:
                            # Non-standard formula
                            new_ws[coordinate].value = None
                            new_ws[coordinate].formula = formula_value
                    except Exception as e:
                        print(f"  Error setting formula in {coordinate}: {e}")
                        new_ws[coordinate].value = formula_value
                else:
                    # Handle non-formula cells
                    if value.lower() == 'true':
                        new_ws[coordinate] = True
                    elif value.lower() == 'false':
                        new_ws[coordinate] = False
                    else:
                        try:
                            if value.isdigit():
                                new_ws[coordinate] = int(value)
                            else:
                                new_ws[coordinate] = float(value)
                        except (ValueError, TypeError):
                            new_ws[coordinate] = value
                
                # Direct copy of all formatting from the source cell
                try:
                    source_cell = source_ws[coordinate]
                    copy_cell_formatting(source_cell, new_ws[coordinate])
                except Exception as e:
                    print(f"  Error copying formatting for {coordinate}: {e}")
        
        # Save the new workbook
        output_file = os.path.join(output_dir, f"{base_name}_recreated.xlsx")
        
        try:
            # Add workbook links for MIS-Report formulas if needed
            if 'Form X Report' in file:
                links_sheet = new_wb.create_sheet(title="_Links", index=0)
                links_sheet["A1"] = "Workbook Index References"
                links_sheet["A2"] = "[1] = Deposits Data Lite.xlsx"
                links_sheet["A3"] = "[2] = Loans Data Lite.xlsx"
                links_sheet["A4"] = "[3] = Form X Report  Main Lite.xlsx"
                links_sheet["A6"] = "Note: These links help resolve formulas with [1], [2] references."
                links_sheet["A7"] = "You may need to update links manually in Excel: Data > Edit Links"
            
            new_wb.save(output_file)
            print(f"Created new workbook: {output_file}")
            recreated_files.append(output_file)
        except Exception as e:
            print(f"ERROR saving workbook {output_file}: {str(e)}")
    
    # Close the database connection
    conn.close()
    
    # Clean up source workbooks
    source_workbooks.clear()
    
    return recreated_files

# ============================================================================
# Phase 4: Font Color Correction
# ============================================================================
def fix_workbook_fonts(recreated_files):
    """Fix ONLY font colors in workbooks while preserving all other formatting."""
    print("\n" + "="*70)
    print("PHASE 4: FONT COLOR CORRECTION")
    print("="*70)
    
    fixed_files = []
    
    for file in recreated_files:
        if not os.path.exists(file):
            print(f"File not found: {file}")
            continue
            
        print(f"Processing file: {file}")
        
        # Load the workbook
        wb = load_workbook(file)
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            print(f"  Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Process all cells
            cells_modified = 0
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Skip empty cells
                    if cell.value is None:
                        continue
                    
                    # Get current font properties and make a copy to modify
                    current_font = cell.font
                    
                    try:
                        # Make exact copy of current font but change color to black
                        new_font = copy(current_font)
                        new_font.color = "FF000000"  # Black with full opacity
                        cell.font = new_font
                        cells_modified += 1
                    except Exception as e:
                        print(f"    Error fixing font in cell {cell.coordinate}: {e}")
            
            print(f"    Modified {cells_modified} cells in {sheet_name}")
        
        # Create output filename
        base_name, ext = os.path.splitext(file)
        output_file = f"{base_name}_fixed{ext}"
        
        # Save the modified workbook
        wb.save(output_file)
        print(f"Saved fixed file: {output_file}")
        fixed_files.append(output_file)
    
    return fixed_files

# ============================================================================
# Main Function
# ============================================================================
def main():
    """Main function to execute the entire Excel processing workflow."""
    global new_base_path
    
    print("\n" + "="*70)
    print("EXCEL PROCESSING SYSTEM")
    print("="*70)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Get the new base path for external references
    new_base_path = input("Enter new base path for external references (leave empty to keep original): ")
    if new_base_path and not new_base_path.endswith('\\'):
        new_base_path += '\\'  # Ensure path ends with backslash
    
    print(f"External reference path will be updated to: '{new_base_path}'" if new_base_path else "External references will keep original paths")
    
    # Make output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")
        
    # Check if the Excel files actually exist
    for file in excel_files:
        if not os.path.exists(file):
            print(f"WARNING: Input file '{file}' does not exist.")
            confirm = input("Continue anyway? (y/n): ")
            if confirm.lower() != 'y':
                print("Process aborted.")
                return
    
    try:
        # Phase 1: Data Identification
        workbook_data = identify_data()
        
        # Phase 2: Data Storage
        store_data(workbook_data)
        
        # Phase 3: Data Recreation with Direct Formatting Copy
        recreated_files = recreate_workbooks()
        
        # Phase 4: Font Color Correction
        fixed_files = fix_workbook_fonts(recreated_files)
        
        # Final Summary
        print("\n" + "="*70)
        print("PROCESS COMPLETED SUCCESSFULLY")
        print("="*70)
        print(f"Input Files: {len(excel_files)}")
        print(f"Recreated Files: {len(recreated_files)}")
        print(f"Fixed Files: {len(fixed_files)}")
    
    except Exception as e:
        print("\n" + "="*70)
        print("ERROR ENCOUNTERED")
        print("="*70)
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        print("\nProcess terminated with errors.")

if __name__ == "__main__":
    main()